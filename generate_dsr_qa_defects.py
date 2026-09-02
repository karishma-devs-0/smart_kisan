"""
DSR covering the work after the 28-29 Aug report.

  Monday 31 Aug    - onboarding location, email sign-in codes, app stability
  Tuesday 1 Sep    - mail delivery on the host, phone codes, profile defect
  Wednesday 2 Sep  - password reset, and the pump/dashboard defects

Same layout as the previous DSR generators. Every row corresponds to a real
commit on main.
"""

import os
import sys
import subprocess

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MON = "2026-08-31"
TUE = "2026-09-01"
WED = "2026-09-02"
FILENAME = "SmartKisan_DSR_2026-08-31_to_09-02.xlsx"

HEADERS = ["Date", "Task Category", "Detailed Task Description",
           "Deliverables / Artifacts", "Status", "Hours Spent", "Remarks & Notes"]

ROWS = [
    # ── Monday 31 Aug ────────────────────────────────────────────────────────
    [
        MON,
        "Onboarding - Farm Location",
        "Onboarding never asked where the farm is, so the app fell back to a "
        "default set of coordinates. The weather screens and the farm map "
        "widget both read from that, which is why testers saw weather for "
        "somewhere other than their farm. Setup now collects the location and "
        "the app applies it on every launch.",
        "OnboardingScreen.js, settingsSlice.js, App.js",
        "Fixed",
        "",
        "Addresses TC-014. The location is stored on the farm profile, so it "
        "survives reinstalling the app.",
    ],
    [
        MON,
        "Authentication - One-Time Codes",
        "Built one-time codes on the server for three flows that had no "
        "implementation: signing in without a password, confirming an email "
        "address at sign-up, and resetting a forgotten password. Codes are "
        "stored hashed, expire in ten minutes, allow five attempts and can be "
        "used once.",
        "otpService.js, mailService.js, auth.js, addAuthCodesSchema.js",
        "Completed",
        "",
        "A request answers identically whether or not the address is "
        "registered, so the endpoint cannot be used to discover which emails "
        "have accounts.",
    ],
    [
        MON,
        "Login - Replacing Dead Tabs",
        "The Phone and Username login tabs could not work: there was no SMS "
        "provider and no username lookup, so both reported themselves "
        "unavailable after the user had filled the form. Replaced with email "
        "code sign-in, which gives a genuine second way into an account.",
        "LoginScreen.js, CodeLoginForm.js, LoginTabBar.js",
        "Fixed",
        "",
        "Addresses TC-007 and TC-008. Also fixes the reported problem of an "
        "error from one tab staying on screen after switching to another.",
    ],
    [
        MON,
        "Stability - Expo Go Crash",
        "The weed detection model loader was imported at start-up. It is a "
        "native module, so on Expo Go the app crashed before reaching the "
        "login screen. The loader is now resolved only when a scan is run.",
        "weedInference.js",
        "Fixed",
        "",
        "Made the app testable in Expo Go again; the model itself still needs a "
        "development build.",
    ],
    [
        MON,
        "Stability - Blank Screen at Launch",
        "While the session was being restored the app rendered nothing. On the "
        "free hosting tier the server can take about fifty seconds to wake, so "
        "a cold start showed a white screen for the whole period and looked "
        "like a hang. Replaced with a loading indicator.",
        "App.js, RootNavigator.js",
        "Fixed",
        "",
        "Defect introduced by our own earlier change; found in testing.",
    ],
    [
        MON,
        "Deployment - Mail Configuration",
        "Declared the mail credentials in the deployment blueprint. The local "
        "environment file is not committed, so the hosted service had none of "
        "them and code delivery failed there while working locally.",
        "render.yaml, .env.example",
        "Completed",
        "",
        "Secrets remain outside the repository; the blueprint only declares "
        "that they are required.",
    ],

    # ── Tuesday 1 Sep ────────────────────────────────────────────────────────
    [
        TUE,
        "Email Delivery - Host Restriction",
        "Sign-in codes hung for over two minutes on the hosted service while "
        "the health check answered in under a second. The cause is that the "
        "host blocks outbound SMTP to prevent spam, and the connection stalls "
        "rather than failing. Moved delivery to an HTTPS mail API, which is "
        "unaffected, keeping SMTP for local development.",
        "mailService.js, render.yaml",
        "Fixed",
        "",
        "Reported as the code request spinning forever and then saying the "
        "request had been cancelled. Connection timeouts were also added so a "
        "future block fails quickly instead of hanging.",
    ],
    [
        TUE,
        "Authentication - Phone Codes",
        "Extended one-time codes to phone numbers, since many farmers do not "
        "have an email address. One code path serves both channels; only "
        "delivery differs. Numbers are normalised so that +91, a leading zero "
        "and spacing all resolve to the same account.",
        "smsService.js, otpService.js, addPhoneAuthSchema.js",
        "Built - Not Live",
        "",
        "Cannot deliver until DLT registration with TRAI is complete. No "
        "provider will send to an Indian number before then. This is a "
        "business registration, not development work.",
    ],
    [
        TUE,
        "Profile - Update Defect",
        "The profile screen read the farm name and location from the user "
        "record, which holds only id, name and email, so both fields opened "
        "blank however much had been entered during setup. They belong to the "
        "farm profile and are now read from there. Saving also did nothing at "
        "all: the update was applied to local state and discarded on restart. "
        "Added the missing server endpoint and the missing validation.",
        "UserProfileScreen.js, authSlice.js, api.js, auth.js (PUT /auth/me)",
        "Fixed",
        "",
        "Addresses TC-019. The reported symptom was the blank fields; the "
        "silent failure to save was found while fixing it. Phone numbers are "
        "now unique across accounts.",
    ],

    # ── Wednesday 2 Sep ──────────────────────────────────────────────────────
    [
        WED,
        "Authentication - Password Reset",
        "Added the password reset screen. The server could already issue and "
        "verify reset codes and the app had the logic wired to them, but "
        "nothing reached it: the 'Forgot password?' link on the login form had "
        "no action attached. A user who forgot their password had no way back "
        "into their account.",
        "ForgotPasswordScreen.js, AuthStack.js, EmailLoginForm.js",
        "Fixed",
        "",
        "Verified by an automated check covering the refusals, the reset "
        "itself, the old password ceasing to work and a code not being "
        "reusable. Thirteen checks, all passing.",
    ],
    [
        WED,
        "Pumps - Queries Against Missing Columns",
        "Four database queries named columns that do not exist, so each failed "
        "every time it ran. Adding a pump always returned an error, which is "
        "why a new pump appeared in the list and was gone after restarting. "
        "Editing a pump had the same fault. Switching a pump on or off "
        "returned an error after having already switched it. The real-time "
        "handler failed at its first query.",
        "pumps.js, mqttService.js",
        "Fixed",
        "",
        "Addresses TC-012. Not one pump run had ever been recorded as a "
        "result: the history table was empty. History writes were also moved "
        "out of the path that can fail the request, since losing a log line "
        "must not fail the action it describes.",
    ],
    [
        WED,
        "Dashboard - Invented Figures",
        "The Today card multiplied the number of running pumps by fixed "
        "constants to produce run hours, water used and power used. This is "
        "why every login showed identical numbers. With runs now recorded, the "
        "figures are derived from real data: litres from each pump's flow "
        "rate and kilowatt-hours from its horsepower, against the time it "
        "actually ran, over the farm's own day rather than the server's.",
        "pumps.js (GET /pumps/summary/today), HomeScreen.js, pumpsSlice.js",
        "Fixed",
        "",
        "Addresses TC-010. The card shows a dash rather than a zero when "
        "nothing has run, so it never states a total it does not have.",
    ],
    [
        WED,
        "Verification - Automated API Check",
        "Extended the end-to-end check to cover the profile update and the "
        "full pump chain: create a pump, run it, confirm the run was recorded "
        "and confirm the derived totals are arithmetically correct.",
        "backend/scripts/e2eCheck.js, backend/scripts/checkPasswordReset.js",
        "Completed",
        "",
        "55 of 55 checks passing. The pump defects above were found by writing "
        "this, not by reading the code.",
    ],

    # ── In progress ──────────────────────────────────────────────────────────
    [
        WED,
        "Crop Suitability - Real Inputs",
        "The recommendation engine is real and works, but is being fed sample "
        "soil and climate figures rather than the farm's own. Wiring it to the "
        "recorded soil readings and the farm's location.",
        "cropRecommendEngine.js, api.js",
        "In Progress",
        "",
        "Addresses TC-015. The climate half also needs the weather API key, "
        "which is not yet configured.",
    ],
    [
        WED,
        "Farm Management - No Data Source",
        "The farm management screen shows sample tasks because there is no "
        "table or endpoint behind it. Adding the storage and the endpoints so "
        "the screen reflects the farm's own tasks.",
        "New: farm_tasks table, backend/src/routes/farmTasks.js",
        "In Progress",
        "",
        "Addresses TC-020. This is new work rather than a defect fix, as the "
        "feature was never connected to anything.",
    ],
]

OPEN_ITEMS = [
    ["TC-016", "Disease detection reported as showing a rejection",
     "Awaiting the exact error text from the testing team. The most likely "
     "cause is the model service waking from idle, which takes a few seconds "
     "on the free tier."],
    ["Phone sign-in", "Codes cannot be delivered to Indian numbers",
     "Requires DLT registration with TRAI before any provider will send. "
     "Business registration, not development."],
    ["Weather", "Forecasts are sample data",
     "Needs an OpenWeatherMap API key. A free key resolves it; the code is "
     "already in place and falls back only because the key is absent."],
    ["Field monitor", "Weed detection needs a development build",
     "The model is a native module and cannot run in Expo Go, so it can only "
     "be confirmed on an installed build."],
]

HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="C8E6C9")
THIN = Side(style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_COLOURS = {
    "Fixed": "1B5E20",
    "Completed": "1B5E20",
    "In Progress": "E65100",
    "Built - Not Live": "E65100",
}


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "DSR"

    ws.append(["SmartKisan - Daily Status Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1B5E20")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.append(["Period: 31 August - 2 September 2026 (following the 28-29 August report)"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1).font = Font(italic=True, size=10, color="555555")

    ws.append([])

    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=4, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row in ROWS:
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status_cell = ws.cell(row=r, column=5)
        colour = STATUS_COLOURS.get(row[4])
        if colour:
            status_cell.font = Font(bold=True, color=colour)

    ws.append([])
    ws.append(["Open Items"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
    ws.cell(row=r, column=1).font = Font(bold=True, size=12, color="1B5E20")
    ws.cell(row=r, column=1).fill = SECTION_FILL

    ws.append(["Reference", "Item", "Blocked On / Next Step", "", "", "", ""])
    r = ws.max_row
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F5E9")
        cell.border = BORDER

    for item in OPEN_ITEMS:
        ws.append([item[0], item[1], item[2], "", "", "", ""])
        r = ws.max_row
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [12, 26, 62, 34, 16, 12, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), FILENAME)
    wb.save(out)
    print("Wrote " + out)
    print("  %d task rows, %d open items" % (len(ROWS), len(OPEN_ITEMS)))


if __name__ == "__main__":
    build()
